__author__ = 'daniele'

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter


import argparse

# PARSER PARAMETRI

parser = argparse.ArgumentParser(description="Extract text from cells in a given column and remove all the words of a stop world list")

parser.add_argument("file_in", help="specify file in")
parser.add_argument("column", help="specify column to elaborate")
parser.add_argument("-o", "--file_out", help="specify file out, if missing will be file-cleared")
parser.add_argument("-e", "--elaborate", help="force elaborate data, otherwise it's just a preview", action="store_true")
parser.add_argument("-p", "--pipe_exp", help="write a pipe between words", action="store_true")

args = parser.parse_args()

#obbligatori

filename_in = args.file_in
column=args.column.upper()

#opzionali

if args.file_out:
    filename_out = args.file_out
else:
	filename_out = "file-cleared.xlsx"

if args.elaborate:
    elaborate = args.elaborate
else:
	elaborate = False

if args.pipe_exp:
    pipe = args.pipe_exp
else:
	pipe = False


# init oggetti excel

wb = load_workbook(filename_in)
ws = wb.active

print("- FILE " + filename_in + " || loaded")
for sheet in wb:
	print("- SHEET " + str(sheet.title)+ " || loaded")

# Carica stop-list
# NOTA: ripendare questa parte caricando dei file esterni

PUNCTUACTION = [".",",",";",":","-","!","?","'","1","2","3","4","5","6","7","8","9","0","/"]
DARKLIST = ["rt","ciao","adoro","soprattutto","spingono","ama","prima","amo","vai","poter","dia","preferisco","con","chi","ce"]
IT = ["IE","a","abbastanza","abbia","abbiamo","abbiano","abbiate","accidenti","ad","adesso","affinche","agl","agli","ahime","ahimè","ai","al","alcuna","alcuni","alcuno","all","alla","alle","allo","allora","altri","altrimenti","altro","altrove","altrui","anche","ancora","anni","anno","ansa","anticipo","assai","attesa","attraverso","avanti","avemmo","avendo","avente","aver","avere","averlo","avesse","avessero","avessi","avessimo","aveste","avesti","avete","aveva","avevamo","avevano","avevate","avevi","avevo","avrai","avranno","avrebbe","avrebbero","avrei","avremmo","avremo","avreste","avresti","avrete","avrà","avrò","avuta","avute","avuti","avuto","basta","bene","benissimo","berlusconi","brava","bravo","c","casa","caso","cento","certa","certe","certi","certo","che","chi","chicchessia","chiunque","ci","ciascuna","ciascuno","cima","cio","cioe","cioè","circa","citta","città","ciò","co","codesta","codesti","codesto","cogli","coi","col","colei","coll","coloro","colui","come","cominci","comunque","con","concernente","conciliarsi","conclusione","consiglio","contro","cortesia","cos","cosa","cosi","così","cui","d","da","dagl","dagli","dai","dal","dall","dalla","dalle","dallo","dappertutto","davanti","degl","degli","dei","del","dell","della","delle","dello","dentro","detto","deve","di","dice","dietro","dire","dirimpetto","diventa","diventare","diventato","dopo","dov","dove","dovra","dovrà","dovunque","due","dunque","durante","e","ebbe","ebbero","ebbi","ecc","ecco","ed","effettivamente","egli","ella","entrambi","eppure","era","erano","eravamo","eravate","eri","ero","esempio","esse","essendo","esser","essere","essi","ex","fa","faccia","facciamo","facciano","facciate","faccio","facemmo","facendo","facesse","facessero","facessi","facessimo","faceste","facesti","faceva","facevamo","facevano","facevate","facevi","facevo","fai","fanno","farai","faranno","fare","farebbe","farebbero","farei","faremmo","faremo","fareste","faresti","farete","farà","farò","fatto","favore","fece","fecero","feci","fin","finalmente","finche","fine","fino","forse","forza","fosse","fossero","fossi","fossimo","foste","fosti","fra","frattempo","fu","fui","fummo","fuori","furono","futuro","generale","gia","giacche","giorni","giorno","già","gli","gliela","gliele","glieli","glielo","gliene","governo","grande","grazie","gruppo","ha","haha","hai","hanno","ho","i","ieri","il","improvviso","in","inc","infatti","inoltre","insieme","intanto","intorno","invece","io","l","la","lasciato","lato","lavoro","le","lei","li","lo","lontano","loro","lui","lungo","luogo","là","ma","macche","magari","maggior","mai","male","malgrado","malissimo","mancanza","marche","me","medesimo","mediante","meglio","meno","mentre","mesi","mezzo","mi","mia","mie","miei","mila","miliardi","milioni","minimi","ministro","mio","modo","molti","moltissimo","molto","momento","mondo","mosto","nazionale","ne","negl","negli","nei","nel","nell","nella","nelle","nello","nemmeno","neppure","nessun","nessuna","nessuno","niente","no","noi","non","nondimeno","nonostante","nonsia","nostra","nostre","nostri","nostro","novanta","nove","nulla","nuovo","o","od","oggi","ogni","ognuna","ognuno","oltre","oppure","ora","ore","osi","ossia","ottanta","otto","paese","parecchi","parecchie","parecchio","parte","partendo","peccato","peggio","per","perche","perchè","perché","percio","perciò","perfino","pero","persino","persone","però","piedi","pieno","piglia","piu","piuttosto","più","po","pochissimo","poco","poi","poiche","possa","possedere","posteriore","posto","potrebbe","preferibilmente","presa","press","prima","primo","principalmente","probabilmente","proprio","puo","pure","purtroppo","può","qualche","qualcosa","qualcuna","qualcuno","quale","quali","qualunque","quando","quanta","quante","quanti","quanto","quantunque","quasi","quattro","quel","quella","quelle","quelli","quello","quest","questa","queste","questi","questo","qui","quindi","realmente","recente","recentemente","registrazione","relativo","riecco","salvo","sara","sarai","saranno","sarebbe","sarebbero","sarei","saremmo","saremo","sareste","saresti","sarete","sarà","sarò","scola","scopo","scorso","se","secondo","seguente","seguito","sei","sembra","sembrare","sembrato","sembri","sempre","senza","sette","si","sia","siamo","siano","siate","siete","sig","solito","solo","soltanto","sono","sopra","sotto","spesso","srl","sta","stai","stando","stanno","starai","staranno","starebbe","starebbero","starei","staremmo","staremo","stareste","staresti","starete","starà","starò","stata","state","stati","stato","stava","stavamo","stavano","stavate","stavi","stavo","stemmo","stessa","stesse","stessero","stessi","stessimo","stesso","steste","stesti","stette","stettero","stetti","stia","stiamo","stiano","stiate","sto","su","sua","subito","successivamente","successivo","sue","sugl","sugli","sui","sul","sull","sulla","sulle","sullo","suo","suoi","tale","tali","talvolta","tanto","te","tempo","ti","titolo","torino","tra","tranne","tre","trenta","troppo","trovato","tu","tua","tue","tuo","tuoi","tutta","tuttavia","tutte","tutti","tutto","uguali","ulteriore","ultimo","un","una","uno","uomo","va","vale","vari","varia","varie","vario","verso","vi","via","vicino","visto","vita","voi","volta","volte","vostra","vostre","vostri","vostro","è"]
EN = ["a","a's","able","about","above","according","accordingly","across","actually","after","afterwards","again","against","ain't","all","allow","allows","almost","alone","along","already","also","although","always","am","among","amongst","an","and","another","any","anybody","anyhow","anyone","anything","anyway","anyways","anywhere","apart","appear","appreciate","appropriate","are","aren't","around","as","aside","ask","asking","associated","at","available","away","awfully","b","be","became","because","become","becomes","becoming","been","before","beforehand","behind","being","believe","below","beside","besides","best","better","between","beyond","both","brief","but","by","c","c'mon","c's","came","can","can't","cannot","cant","cause","causes","certain","certainly","changes","clearly","co","com","come","comes","concerning","consequently","consider","considering","contain","containing","contains","corresponding","could","couldn't","course","currently","d","definitely","described","despite","did","didn't","different","do","does","doesn't","doing","don't","done","down","downwards","during","e","each","edu","eg","eight","either","else","elsewhere","enough","entirely","especially","et","etc","even","ever","every","everybody","everyone","everything","everywhere","ex","exactly","example","except","f","far","few","fifth","first","five","followed","following","follows","for","former","formerly","forth","four","from","further","furthermore","g","get","gets","getting","given","gives","go","goes","going","gone","got","gotten","greetings","h","had","hadn't","happens","hardly","has","hasn't","have","haven't","having","he","he's","hello","help","hence","her","here","here's","hereafter","hereby","herein","hereupon","hers","herself","hi","him","himself","his","hither","hopefully","how","howbeit","however","i","i'd","i'll","i'm","i've","ie","if","ignored","immediate","in","inasmuch","inc","indeed","indicate","indicated","indicates","inner","insofar","instead","into","inward","is","isn't","it","it'd","it'll","it's","its","itself","j","just","k","keep","keeps","kept","know","known","knows","l","last","lately","later","latter","latterly","least","less","lest","let","let's","like","liked","likely","little","look","looking","looks","ltd","m","mainly","many","may","maybe","me","mean","meanwhile","merely","might","more","moreover","most","mostly","much","must","my","myself","n","name","namely","nd","near","nearly","necessary","need","needs","neither","never","nevertheless","new","next","nine","no","nobody","non","none","noone","nor","normally","not","nothing","novel","now","nowhere","o","obviously","of","off","often","oh","ok","okay","old","on","once","one","ones","only","onto","or","other","others","otherwise","ought","our","ours","ourselves","out","outside","over","overall","own","p","particular","particularly","per","perhaps","placed","please","plus","possible","presumably","probably","provides","q","que","quite","qv","r","rather","rd","re","really","reasonably","regarding","regardless","regards","relatively","respectively","right","s","said","same","saw","say","saying","says","second","secondly","see","seeing","seem","seemed","seeming","seems","seen","self","selves","sensible","sent","serious","seriously","seven","several","shall","she","should","shouldn't","since","six","so","some","somebody","somehow","someone","something","sometime","sometimes","somewhat","somewhere","soon","sorry","specified","specify","specifying","still","sub","such","sup","sure","t","t's","take","taken","tell","tends","th","than","thank","thanks","thanx","that","that's","thats","the","their","theirs","them","themselves","then","thence","there","there's","thereafter","thereby","therefore","therein","theres","thereupon","these","they","they'd","they'll","they're","they've","think","third","this","thorough","thoroughly","those","though","three","through","throughout","thru","thus","to","together","too","took","toward","towards","tried","tries","truly","try","trying","twice","two","u","un","under","unfortunately","unless","unlikely","until","unto","up","upon","us","use","used","useful","uses","using","usually","uucp","v","value","various","very","via","viz","vs","w","want","wants","was","wasn't","way","we","we'd","we'll","we're","we've","welcome","well","went","were","weren't","what","what's","whatever","when","whence","whenever","where","where's","whereafter","whereas","whereby","wherein","whereupon","wherever","whether","which","while","whither","who","who's","whoever","whole","whom","whose","why","will","willing","wish","with","within","without","won't","wonder","would","wouldn't","x","y","yes","yet","you","you'd","you'll","you're","you've","your","yours","yourself","yourselves","z","zero"]
wlist= IT + EN + DARKLIST
print("- WORDS IT/EN stop-word-list || loaded")


# IN: string of words
# FN: remove punctuaction
# OUT: string cleared

def clear(strData, lista):
	for item in lista:
		strData=strData.replace(item, " ")

	return strData


# IN: list of words
# FN: remove any that are in a list of stop words.
# OUT: an object LIST of words

def removeSW(strData, listaParole):
	n=0
	strData=strData.split()
	for d_item in strData:
		n=n+1
		d_item = d_item.lower()
		# print(str(n) + " " + d_item)
		for w_item in listaParole:
			if d_item==w_item.lower():
				strData[n-1]=""
				# print(strData)
	return(strData)

# IN: list of words
# FN: join words in a single string
# OUT: string

def combine(lista):
	stringa=""
	for item in lista:
		item.strip()
		if item != "":
			stringa=stringa + " " + str(item)

	return stringa

def combinePipe(lista):
	stringa=""
	for item in lista:
		item.strip()
		if item != "":
			stringa=stringa + " | " + str(item)

	return stringa



# MAIN

# clear given column
print("- WAIT FOR", end=' ')
for cell in ws[column]:
	print("", end='.')
	cell.value = str(cell.value)
	if cell.value:
		cell.value = clear(cell.value, PUNCTUACTION)
		if pipe:
			cell.value = str(combinePipe(removeSW(cell.value, wlist)))
		else:
			cell.value = str(combine(removeSW(cell.value, wlist)))
		


# elaborate: se il flag elaborate è true allora elabora il file, altrimenti stampa una preview

if elaborate:
	wb.save(filename_out)
	print('\n')
	print("- FILE elaborated and ready")
else:
	print("- start elaborationa")
	print("-----------------------------------")
	for cell in ws[column]:
		print(cell.value)
		print("-----")
	print("-----------------------------------")
	print("- end elaborationa")















